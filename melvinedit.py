import cv2
import numpy as np
import mediapipe as mp
import tensorflow as tf
from tensorflow.keras.models import load_model
import pygame
import random
from playsound import playsound
import time
import openpyxl
import os


# Initialize pygame
pygame.init()

# Set up the display in full-screen mode
screen = pygame.display.set_mode((1600, 900), pygame.FULLSCREEN)
pygame.display.set_caption("Hand Gesture Game")
# Get the dimensions of the screen
screen_width, screen_height = pygame.display.get_surface().get_size()
fullscreen = True
font = pygame.font.Font(None, 32)
# Set up the text input box
input_box = pygame.Rect(screen_width // 2 - 190, screen_height // 2.3 - 16, 200, 50)
color_inactive = pygame.Color('lightskyblue3')
color_active = pygame.Color('dodgerblue2')
color = color_inactive
text = ''
active = False

button_rect = pygame.Rect(screen_width - 80, screen_height -36, 200, 50)
button_color = pygame.Color('gray')
button_text = font.render('Open Excel', True, pygame.Color('white'))

game_started = False
n=0
s=0
over=0
correct =0
# Load a background image for the start screen
start_screen_bg = pygame.image.load("start_screen_bg.jpg")
start_screen_bg = pygame.transform.scale(start_screen_bg, (screen_width, screen_height))


# Function to display the start screen
def display_start_screen():
    # Display the background image
    screen.blit(start_screen_bg, (0, 0))
    
    font = pygame.font.Font(None, 60)

    # Style the "Press Enter to Start" text
    start_font = pygame.font.Font(None, 72)
    text_color = (255, 255, 255)
    text_shadow_color = (0, 0, 0)
    start_text = start_font.render("Press Enter to Start", True, text_color)
    
    # Create a shadow for the text
    shadow_offset = 4
    shadow_text = start_font.render("Press Enter to Start", True, text_shadow_color)
    
    # Display the shadow text
    screen.blit(shadow_text, (screen_width // 2 - shadow_text.get_width() // 2 + shadow_offset,
                              screen_height // 2 - shadow_text.get_height() // 2 + shadow_offset))
    
    # Display the actual text
    screen.blit(start_text, (screen_width // 2 - start_text.get_width() // 2,
                            screen_height // 2 - start_text.get_height() // 2))

    # Draw the input box with white border
    pygame.draw.rect(screen, (255, 255, 255), input_box, 2)
    
    # Render the text and display it in the input box
    text_surface = font.render(text, True, color)
    screen.blit(text_surface, (input_box.x + 5, input_box.y + 5))
    input_box.w = max(400, text_surface.get_width() + 10)

    pygame.draw.rect(screen, (0, 255, 0), button_rect)
    font = pygame.font.Font(None, 24)
    text1 = font.render("details", True, (255, 255, 255))
    screen.blit(text1, (button_rect.x + 10, button_rect.y + 15))
    for event in pygame.event.get():
        if event.type == pygame.QUIT:
            running = False
        elif event.type == pygame.MOUSEBUTTONDOWN:
            # Check if the mouse click is within the button rectangle
            if button_rect.collidepoint(event.pos):
                # Open the 'players.xlsx' file
                os.startfile('players.xlsx')

    

    

# Load available gestures and their corresponding images
gesture_images = {
    "rock": pygame.image.load("rock.jpg"),
    "fist": pygame.image.load("fist.jpg"),
    # ['okay', 'peace', 'thumbs up', 'thumbs down', 'call me', 'stop', 'rock', 'live long', 'fist', 'smile']
    "okay": pygame.image.load("ok.jpg"),
    "peace": pygame.image.load("peace.jpg"),
    "thumbs up": pygame.image.load("thumbs up.jpg"),

    "stop": pygame.image.load("stop.jpg"),
    "smile": pygame.image.load("smile.jpg"),
    "live long": pygame.image.load("live long.jpg"),
}

# Create a list of available gesture names
available_gestures = list(gesture_images.keys())

# Initialize game variables
score = 0
user_gestures = []
gesture_read_time = 0  # Timestamp when gesture was last read
gesture_cooldown = 2  # Cooldown time before reading a new gesture (in seconds)
feedback_duration = 3  # Increase the feedback duration to 3 seconds
feedback_start_time = 0  # Timestamp when feedback started
game_over = False  # Game over state

# Initialize feedback variables
feedback_color = (0, 0, 0)  # Default feedback color (black)
feedback_text = ""

# Initialize mediapipe
mp_hands = mp.solutions.hands
hands = mp_hands.Hands(max_num_hands=1, min_detection_confidence=0.7)
mp_draw = mp.solutions.drawing_utils

# Load the gesture recognizer model
model = load_model('mp_hand_gesture')

# Load class names
with open('gesture.names', 'r') as f:
    class_names = f.read().split('\n')

# Initialize the webcam
cap = cv2.VideoCapture(0)

# Flag to track full-screen mode
fullscreen = False
# Timer variables
countdown_start_time = time.time()
countdown_duration = 60 #10 sec
countdown_color = (0, 0, 255)  # Default color (blue)



while True:
    for event in pygame.event.get():
        if event.type == pygame.QUIT:
            running = False
        if event.type == pygame.MOUSEBUTTONDOWN:
            # If the user clicked on the input box, activate it
            if input_box.collidepoint(event.pos):
                active = not active
            else:
                active = False
            # If the user clicked on the button, open the Excel sheet
            if button_rect.collidepoint(event.pos):
                wb = openpyxl.load_workbook('players.xlsx')
                ws = wb.active
                for row in ws.iter_rows(min_row=2, max_col=1):
                    for cell in row:
                        print(cell.value)
                wb.close()
        if event.type == pygame.KEYDOWN:
            # If the user pressed a key, add it to the text input box
            if active:
                if event.key == pygame.K_RETURN:
                    # If the user pressed Enter, store the name in the Excel sheet
                    wb = openpyxl.load_workbook('players.xlsx')
                    ws = wb.active
                    row = ws.max_row + 1
                    ws.cell(row=row, column=1, value=text)
                    wb.save('players.xlsx')
                    wb.close()
                    text = ''
                elif event.key == pygame.K_BACKSPACE:
                    text = text[:-1]
                else:
                    text += event.unicode
            # Change the color of the input box depending on whether it's active or not
            color = color_active if active else color_inactive
                
        # Check for keyboard events
        if event.type == pygame.KEYDOWN:
            if event.key == pygame.K_ESCAPE:
                # Exit full-screen mode
                pygame.quit()
                cap.release()
                cv2.destroyAllWindows()
                exit()
            elif event.key == pygame.K_f:
                if fullscreen:
                    # Exit full-screen mode
                    pygame.display.set_mode((800, 600))
                    screen = pygame.display.set_mode((800, 600))
                    pygame.display.set_caption("Hand Gesture Game")
                else:
                    # Enter full-screen mode
                    screen = pygame.display.set_mode((800, 600))
                    fullscreen = True
            elif event.key == pygame.K_RETURN:
                if not game_started:
                    game_started = True
                    initial_game_start = True  # Mark initial game start
                    show_countdown_timer = True 
                    

    # Display the start screen if the game has not started
    if not game_started:
        display_start_screen()
        pygame.display.update() 
    else:
            

            # Calculate the feedback elapsed time
            current_time = time.time()
            feedback_elapsed_time = current_time - feedback_start_time
            
            if fullscreen == False:
                # Read each frame from the webcam       
                _, frame = cap.read()

                # Flip the frame vertically
                frame = cv2.flip(frame, 1)
                frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

                # Get hand landmark prediction
                result = hands.process(frame)

                className = ''
            

                frame = cv2.resize(frame, (screen_width, screen_height))

                x, y, _ = frame.shape
            else:
                # Read each frame from the webcam
                _, frame = cap.read()

                # Flip the frame vertically
                frame = cv2.flip(frame, 1)
                frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

                # Get hand landmark prediction
                result = hands.process(frame)

                className = ''

                frame = cv2.resize(frame, (800, 600))

                x, y, _ = frame.shape

            # Post-process the result
            if result.multi_hand_landmarks:
                current_time = time.time()
                elapsed_time = current_time - gesture_read_time
                landmarks = []
                for handslms in result.multi_hand_landmarks:
                    for lm in handslms.landmark:
                        lmx = int(lm.x * x)
                        lmy = int(lm.y * y)
                        landmarks.append([lmx, lmy])
                    # Drawing landmarks on frames
                    mp_draw.draw_landmarks(frame, handslms, mp_hands.HAND_CONNECTIONS)
                    # Predict gesture
                    prediction = model.predict([landmarks], verbose=0)
                    classID = np.argmax(prediction)
                    className = class_names[classID]
                    gesture_read_time = current_time

            # disply video
            surface = pygame.surfarray.make_surface(frame)
            surface = pygame.transform.rotate(surface, -90)
            screen.blit(surface, (0, 0))

            # Show the prediction on the frame
            pred_font = pygame.font.Font(None, 48)
            pred_font_render = pred_font.render(className, True, (255, 255, 255))
            screen.blit(pred_font_render, (20, 600 - pred_font_render.get_height() - 20))

            # Get the current target gesture
            current_target = available_gestures[score % len(available_gestures)]
            target_image = gesture_images[current_target]
            if n==0:
                target_image = pygame.transform.scale(target_image, (screen_width, screen_height))
                screen.blit(target_image, (0, 0))
            
                pygame.display.update()
                
            elif n==1:
                target_image = pygame.transform.scale(target_image, (300, 300))
                screen.blit(target_image, (0, 0))
                pygame.display.update()
            # Show the cv2 preview on the screen
            
                

            # Show feedback animation
            current_time = time.time()
            feedback_elapsed_time = current_time - feedback_start_time
            if feedback_elapsed_time <= feedback_duration:
                feedback_alpha = int(255 - (feedback_elapsed_time / feedback_duration) * 255)
                feedback_surface = pygame.Surface((screen_width, screen_height), pygame.SRCALPHA)
                feedback_surface.fill((feedback_color[0], feedback_color[1], feedback_color[2], feedback_alpha))
                screen.blit(feedback_surface, (0, 0))
                font = pygame.font.Font(None, 48)
                feedback_text_render = font.render(feedback_text, True, (255, 255, 255))
                screen.blit(feedback_text_render, (400 - feedback_text_render.get_width() // 2, 300))

            
       
            if className == current_target:
                if className not in user_gestures:
                    n=n+1
                    user_gestures.append(className)
                    score += 1
                    feedback_color = (0, 250, 0)  # Green for correct gesture
                    feedback_text = "Correct!"
                    correct+=1
                    feedback_start_time = current_time  # Reset feedback timer for correct gesture
                    pygame.draw.rect(screen, feedback_color, (0, 0, screen_width, screen_height))
                    if score == 1:
                            game_over = True
                    countdown_start_time = current_time  # Restart the countdown timer
                    countdown_elapsed_time = current_time - countdown_start_time
                    remaining_time= max(0, countdown_duration - countdown_elapsed_time)
                    # Draw the "Correct!" text
                    font = pygame.font.Font(None, 100)
                    text_render = font.render(feedback_text, True, (255, 255, 255))
                    text_rect = text_render.get_rect(center=(650, 300))
                    screen.blit(text_render, text_rect.center)

                    pygame.display.flip()  # Update the display
                    playsound('happysound.mp3') 
                    # Wait for 3 seconds
                    time.sleep(1.5)
                    
                    feedback_color = (0, 0, 0)
                    feedback_text = ""
                

                    # Start the countdown timer
                    countdown_start_time = current_time
                    countdown_color = (0, 0, 255) 
                   # Reset countdown color to blue
                    if remaining_time <=11:
                        countdown_color = (255,0,0) 
                    n=0
                else:
                    
                    feedback_color = (0, 0, 0)
                    feedback_text = ""
            
            else:
                
                n=1
                countdown_elapsed_time = current_time - countdown_start_time
                remaining_time= max(0, countdown_duration - countdown_elapsed_time)
                if remaining_time <= 0:
                    user_gestures.append(className)
                    feedback_color = (255, 0, 0)  # Green for correct gesture
                    feedback_text = "Oh sorry, Time is over"
                    over+=1
                    n=1
                    if(over == 4):
                        game_over = True
                    feedback_start_time = current_time  # Reset feedback timer for correct gesture
                    pygame.draw.rect(screen, feedback_color, (0, 0, screen_width, screen_height))
                    
                    # Draw the "Correct!" text
                    font = pygame.font.Font(None, 100)
                    text_render = font.render(feedback_text, True, (255, 255, 255))
                    text_rect = text_render.get_rect(center=(450, 300))
                    screen.blit(text_render, text_rect.center)
                    pygame.display.flip()  # Update the display
                    playsound('timeout.mp3') 
                    # Wait for 3 seconds
                    time.sleep(1)
                    feedback_color = (0, 0, 0)
                    feedback_text = ""

                    # Start the countdown timer
                    countdown_start_time = current_time
                    countdown_color = (0, 0, 255) 
                    # Reset countdown color to blue
                    if remaining_time <=11:
                        countdown_color = (255,0,0) 
            
                        
                font = pygame.font.Font(None, 72)
                countdown_text = font.render(f"Time Left: {int(remaining_time)}", True, countdown_color)
                screen.blit(countdown_text, (screen_width // 2 - countdown_text.get_width() // 2, 20))
                                
                if game_over :
                    #display game over screen and final score
                    pygame.display.update()
                    screen.blit(start_screen_bg, (0, 0))
                    font = pygame.font.Font(None, 72)
                    text = font.render(f"Game Over! Your score is {score} press enter to restart", True, (255, 255, 255))
                    screen.blit(text, (screen_width // 2 - text.get_width() // 2, screen_height // 2 - text.get_height() // 2))
                    pygame.display.flip()
                    playsound('endsound.mp3')
                    keys = pygame.key.get_pressed()
                    if keys[pygame.K_RETURN]:  # Check if the Enter key is pressed
                        # Reset game variables
                        score = 0
                        user_gestures = []
                        game_over = False
                        restart_pending = False
                                    

            # Draw score on the screen
    font = pygame.font.Font(None, 36)
    score_text = font.render(f"Score: {score}", True, (255, 255, 255))
    screen.blit(score_text, (screen_width - score_text.get_width() - 10, 0))

    # Update the Pygame display
    pygame.display.update()

    if cv2.waitKey(1) == ord('q'):
        break

# Release the webcam and destroy all active windows
cap.release()
cv2.destroyAllWindows()